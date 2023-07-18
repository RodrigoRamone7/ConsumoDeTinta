import openpyxl as xl
from openpyxl.styles import numbers
from datetime import datetime

def gravarDados(m,c,a,b,wh1,wh2):
    def data_hora():
        # Obter a data e hora atuais
        data_hora_atual = datetime.now()

        # Formatando a data e hora no formato desejado
        formato = "%d/%m/%Y %H:%M"
        data_hora_formatada = data_hora_atual.strftime(formato)
        return data_hora_formatada

    # Carregar o arquivo da planilha existente
    wb = xl.load_workbook('Consumo.xlsx')
    wb.save('Backup.xlsx')

    # Escolhe active sheet
    ws = wb.active

    # Encontrar a próxima célula vazia na coluna A
    next_row = 1
    while ws.cell(row=next_row, column=1).value is not None:
        next_row += 1

    magenta = m
    cyan = c
    amarelo = a
    preto = b
    branco1 = wh1
    branco2 = wh2
    
    def limparcaractere(limp):
        l = limp
        limpar = "%\n"
        for letra in l:
            if letra in limpar:
                l = l.replace(letra,'')
        try:
            decimal = int(l) / 100
            return decimal
        except ValueError:
            return None
    
    magenta = limparcaractere(magenta)
    cyan = limparcaractere(cyan)
    amarelo = limparcaractere(amarelo)
    preto = limparcaractere(preto)
    branco1 = limparcaractere(branco1)
    branco2 = limparcaractere(branco2)
    
    def converterML(decimal):
        mls = decimal * 600 / 100 * 100
        return mls
    
    mlMagenta = converterML(magenta)
    mlCyan = converterML(cyan)
    mlAmarelo = converterML(amarelo)
    mlPreto = converterML(preto)
    mlBranco1 = converterML(branco1)
    mlBranco2 = converterML(branco2)
    mlTotal = mlMagenta + mlCyan + mlAmarelo + mlPreto + mlBranco1 + mlBranco2

    # Dados que você deseja adicionar
    dados = [data_hora(),
            magenta, mlMagenta,
            cyan, mlCyan,
            amarelo, mlAmarelo,
            preto, mlPreto,
            branco1, mlBranco1,
            branco2, mlBranco2,
            mlTotal]

    # adiciona os dados na próxima linha vazia sem sequencia
    for i, dado in enumerate(dados, start=1):
        coluna = ws.cell(row=next_row, column=i)
        coluna.value = dado
        


    def format_colum_percent(coluna): #Formata a coluna em porcentagem para valores de string terminados em %
        format_num = numbers.FORMAT_PERCENTAGE
        col = coluna
        for celula in ws[col]:
            celula.number_format = format_num
            valor_format = celula.value
            if isinstance(valor_format, str) and valor_format.endswith('%'):
                valor_num = float(valor_format.strip('%')) / 100
                celula.value = valor_num

    format_colum_percent("B")
    format_colum_percent("D")
    format_colum_percent("F")
    format_colum_percent("H")
    format_colum_percent("J")
    format_colum_percent("L")

    def format_colum_ml(coluna): #Formata a coluna de forma personalizada de acordo com célula especificada
        col = ws[coluna]
        format_num = ws['P1'].number_format #Célula a ser capturada a formatação personalizada
        for celula in col:
            celula.number_format = format_num
            
    format_colum_ml("C")
    format_colum_ml("E")
    format_colum_ml("G")
    format_colum_ml("I")
    format_colum_ml("K")
    format_colum_ml("M")
    format_colum_ml("N")

    wb.save('Consumo.xlsx')

